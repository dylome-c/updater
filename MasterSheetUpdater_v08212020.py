#Reading an excel file using Python 
import openpyxl
from openpyxl.styles import PatternFill
import xlrd
import math
import datetime

DEBUG = 0
SpecificPO = 0
  
# To open Workbook 
masterSheetWb = openpyxl.load_workbook("MasterSheet.xlsx")
GTNWb = openpyxl.load_workbook("GTN.xlsx")

masterSheet = masterSheetWb["MasterSheet"]
newBookingsSheet = masterSheetWb["New Bookings"]
updateSheet = masterSheetWb["Updates"]
lookUpSheet = masterSheetWb["Origin Lookup"]
GTNSheet = GTNWb["Sheet1"]

# Update Styles
yellowFill = PatternFill(start_color="FFFF00", fill_type = "solid")
redFill = PatternFill(start_color="FF0000" , fill_type = "solid")
blueFill = PatternFill(start_color="99CCFF" , fill_type = "solid")
greenFill = PatternFill(start_color="90EE90" , fill_type = "solid")



# Mastersheet Update Main Function
# Purpose: FOR EACH RECORD IN masterSheet, FIND IT'S CORRESPONDING RECORD IN GTNSheet, UPDATE ACCORDINGLY (Number Changes, Cargo detail changes, Cancelled)
def updateExistingBookings(masterSheet, GTNSheet):

    # For each row in the mastersheet
    for row in range(3 , masterSheet.max_row + 1):

        # If PO number field (col:15) is empty, continue to next row
        if str(masterSheet.cell(row, 15).value).rstrip() == "None":
            continue

        # Non-empty row, begin with getting the PO and Line-item ID
        PO = str(masterSheet.cell(row, 15).value).rstrip()
        lineItemID = str(masterSheet.cell(row, 18).value).rstrip()

        # Use this status to see if we find a match in the GTN booking reports
        matchStatus = False


        # If DEBUG is ON, print out Mastersheet PO and LineItemID for debugging purposes
        if DEBUG == 1:
            print("PO: ", PO, ", Line Item ID: ", lineItemID)

        
        # Search for PO+Line Item ID in the GTN booking report, for each row in the GTN booking report
        for matchingRow in range(3, GTNSheet.max_row + 1):

            # If "This" row in the GTN booking report is empty, SKIP
            if str(GTNSheet.cell(matchingRow, 14).value).rstrip() == "None":
                continue

            
            # Now we found a non-empty GTN row with the SAME origin, get the PO and LineItemID in "This" GTN row
            GTNPO = str(GTNSheet.cell(matchingRow, 14).value).rstrip()
            GTNLineItemID = str(GTNSheet.cell(matchingRow, 17).value).rstrip()


            # If DEBUG is ON, print out GTNPO and GTNLineItemID for debugging purposes
            if DEBUG == 1:
                print("GTN PO: ", GTNPO, ", GTN Line Item ID: ", GTNLineItemID)

                
            # If MasterSheetPO matches GTNPO AND MasterSheetLineItemID matches GTNLineItemID
            if  GTNPO == PO and GTNLineItemID == lineItemID:

                # IMPORTANT!!! Toggle matchStatus to True
                matchStatus = True

                # If DEBUG is ON, print out alert to indicate a match was found
                if DEBUG == 1:
                    print("Found match! Row in GTN: ", matchingRow)
                    
    
                # Start Comparison of GTNRow and Mastersheet Row Process:
                # 1. We have to identify any CHANGES to the following metrics
                # a. SO number (GTNCol: 6, MastersheetCol: 7) <--- Integer Comparison √
                # b. HOD, ETD, ETA, INDC, ETA INDC: DON'T CHECK because date data formats are unreliable in Excel, ONLY PASTE
                # c. Quantity (GTNCol: 20, MastersheetCol: 21) <--- Integer Comparison √
                # d. Cartons (GTNCol: 21, MastersheetCol: 22) <--- Integer Comparison √
                # e. Total PackUnits (GTNCol: 30, MastersheetCol:31) <--- Integer Comparison √
                # f. Status (GTNCol: 2, MastersheetCol: 3) <--- String Comparison √
                # g. Volume (GTNCol: 22, MastersheetCol: 23) <--- Float Comparison to 3 decimal places √
                # h. Gross Volume (GTNCol: 24, MastersheetCol: 25) <--- Float Comparison to 3 decimal places √
                # i. Weight (GTNCol: 26, MastersheetCol: 27) <--- Float Comparison to 3 decimal places √
                # j: Gross Weight (GTNCol: 28, MastersheetCol: 29) <--- Float Comparison to 3 decimal places √
                # k: Port of Loading (GTNCol: 32, MastersheetCol: 33) <--- String Comparison √
                # l: All Dates (GTNCol: 33, 35, 37, 38) <--- Null/Non-null Comparison √

                # 2. If difference was found from the above:
                # a. Change color to Yellow
                # b. Paste new value
                # c. Write change notes in mastersheet "Exceptions" (MastersheetCol: 2)
                #   i. If notes empty, then simply write XX to YY
                #   ii. Else: Append XX to YY to the notes
                
                # 3. Else:
                # a. If HOD, ETD, ETA, INDC, ETA INDC, Paste over the GTN value to the Mastersheet value
                # b. Colour cell Green


                # Beginning with Shipper Order Status (GTNCol: 2) in the GTN report
                for GTNCol in range(2, 51):

                    # Matching column in mastersheet is GTNCol + 1
                    masterSheetCol = GTNCol + 1


                    # CASE 1f. Status (GTNCol: 2, MastersheetCol: 3)
                    # CASE 1k: Port of Loading (GTNCol: 32, MastersheetCol: 33)
                    if GTNCol == 2 or GTNCol == 32:

                        # Check if the Status in GTN and Mastersheet matches
                        if str(GTNSheet.cell(matchingRow, GTNCol).value).rstrip() == str(masterSheet.cell(row, masterSheetCol).value).rstrip():

                            # No need paste, just colour the cell Green
                            masterSheet.cell(row, masterSheetCol).fill = greenFill


                        # Else, if Status has changed
                        else:
                            
                            # Update the change description FIRST so we know the old and new value

                            # updatedField represents which field has been updated e.g. "Carton"
                            updatedField = str(masterSheet.cell(2, masterSheetCol).value)

                            # First check if the "Exceptions" (MastersheetCol: 2) is empty
                            if str(masterSheet.cell(row, 2).value).rstrip() == "None":

                                # If it is, add the description "[updatedField] changed from [oldValue] to [newValue]"
                                masterSheet.cell(row, 2).value = updatedField + " changed from " + str(masterSheet.cell(row, masterSheetCol).value).rstrip() +" to " + str(GTNSheet.cell(matchingRow, GTNCol).value).rstrip()

                            else:

                                # If it already has changed once, just append to it ", [updatedField] changed from [oldValue] to [newValue]" instead of replacing the whole cell
                                masterSheet.cell(row, 2).value = str(masterSheet.cell(row, 2).value) + (", " + updatedField + " changed from " + str(masterSheet.cell(row, masterSheetCol).value).rstrip() +" to " + str(GTNSheet.cell(matchingRow, GTNCol).value).rstrip())
                            

                            # Paste this value and colour the cell Yellow
                            masterSheet.cell(row, masterSheetCol).value = GTNSheet.cell(matchingRow, GTNCol).value
                            masterSheet.cell(row, masterSheetCol).fill = yellowFill
                            



                    # CASE 1a. SO number (GTNCol: 6, MastersheetCol: 7)
                    # CASE 1c. Quantity (GTNCol: 20, MastersheetCol: 21)
                    # CASE 1d. Cartons (GTNCol: 21, MastersheetCol: 22)
                    # CASE 1e. Total PackUnits (GTNCol: 30, MastersheetCol:31)
                    elif GTNCol == 6 or GTNCol == 20 or GTNCol == 21 or GTNCol == 30:

                        # Check if these integers match
                        tempstr1 = str(GTNSheet.cell(matchingRow, GTNCol).value)[0:str(GTNSheet.cell(matchingRow, GTNCol).value).find(".")] if str(GTNSheet.cell(matchingRow, GTNCol).value).find(".") > 0 else str(GTNSheet.cell(matchingRow, GTNCol).value)
                        tempstr2 = str(masterSheet.cell(row, masterSheetCol).value)[0:str(masterSheet.cell(row, masterSheetCol).value).find(".")] if str(masterSheet.cell(row, masterSheetCol).value).find(".") > 0 else str(masterSheet.cell(row, masterSheetCol).value)

                        
                        if tempstr1 == tempstr2:

                            # No need paste, just colour the cell Green
                            masterSheet.cell(row, masterSheetCol).fill = greenFill


                        # Else, if these integers have changed
                        else:
                            # Update the change description FIRST so we know the old and new value

                            # updatedField represents which field has been updated e.g. "Volume"
                            updatedField = str(masterSheet.cell(2, masterSheetCol).value)

                            # First check if the "Exceptions" (MastersheetCol: 2) is empty
                            if str(masterSheet.cell(row, 2).value).rstrip() == "None":

                                # If it is, add the description "[updatedField] changed from [oldValue] to [newValue]"
                                masterSheet.cell(row, 2).value = updatedField + " changed from " + str(masterSheet.cell(row, masterSheetCol).value).rstrip() +" to " + str(GTNSheet.cell(matchingRow, GTNCol).value).rstrip()

                            else:

                                # If it already has changed once, just append to it ", [updatedField] changed from [oldValue] to [newValue]" instead of replacing the whole cell
                                masterSheet.cell(row, 2).value = str(masterSheet.cell(row, 2).value) + (", " + updatedField + " changed from " + str(masterSheet.cell(row, masterSheetCol).value).rstrip() +" to " + str(GTNSheet.cell(matchingRow, GTNCol).value).rstrip())
                            

                            # Paste this value and colour the cell Yellow
                            masterSheet.cell(row, masterSheetCol).value = GTNSheet.cell(matchingRow, GTNCol).value
                            masterSheet.cell(row, masterSheetCol).fill = yellowFill
                            
                    # CASE 1g. Volume (GTNCol: 22, MastersheetCol: 23)
                    # CASE 1h. Gross Volume (GTNCol: 24, MastersheetCol: 25)
                    # CASE 1i. Weight (GTNCol: 26, MastersheetCol: 27)
                    # CASE 1j: Gross Weight (GTNCol: 28, MastersheetCol: 29)
                    elif GTNCol == 22 or GTNCol == 24 or GTNCol == 26 or GTNCol == 28:

                        # Check if the GTN value and mastersheet value matches to the closest 0.01
                        if math.isclose(float(GTNSheet.cell(matchingRow, GTNCol).value), float(masterSheet.cell(row, masterSheetCol).value), abs_tol=0.01): 

                            # If they match, no need paste, just colour the cell Green
                            masterSheet.cell(row, masterSheetCol).fill = greenFill
                    
                        else:

                            # Update the change description FIRST so we know the old and new value

                            # updatedField represents which field has been updated e.g. "Volume"
                            updatedField = str(masterSheet.cell(2, masterSheetCol).value)

                            # First check if the "Exceptions" (MastersheetCol: 2) is empty
                            if str(masterSheet.cell(row, 2).value).rstrip() == "None":

                                # If it is, add the description "[updatedField] changed from [oldValue] to [newValue]"
                                masterSheet.cell(row, 2).value = updatedField + " changed from " + str(masterSheet.cell(row, masterSheetCol).value).rstrip() +" to " + str(GTNSheet.cell(matchingRow, GTNCol).value).rstrip()

                            else:

                                # If it already has changed once, just append to it ", [updatedField] changed from [oldValue] to [newValue]" instead of replacing the whole cell
                                masterSheet.cell(row, 2).value = str(masterSheet.cell(row, 2).value) + (", " + updatedField + " changed from " + str(masterSheet.cell(row, masterSheetCol).value).rstrip() +" to " + str(GTNSheet.cell(matchingRow, GTNCol).value).rstrip())
                            

                            # Paste this value and colour the cell Yellow
                            masterSheet.cell(row, masterSheetCol).value = GTNSheet.cell(matchingRow, GTNCol).value
                            masterSheet.cell(row, masterSheetCol).fill = yellowFill

                    
                    # CASE 1l: All Dates (GTNCol: 33, 35, 37, 38)
                    elif GTNCol == 33 or GTNCol == 35 or GTNCol == 37 or GTNCol == 38:

                        # Check if mastersheet is non-null and if GTN sheet is Null, in other words added a date
                        if str(masterSheet.cell(row,masterSheetCol).value).strip() == "None" and str(GTNSheet.cell(matchingRow, GTNCol).value).strip() != "None":
                            # Update the change description to signify that dates have been added


                            # First check if the "Exceptions" (MastersheetCol: 2) is empty
                            if str(masterSheet.cell(row, 2).value).rstrip() == "None":

                                # If it is, cell should be "Dates Added"
                                masterSheet.cell(row, 2).value = "Dates Added"
                            else:
                                # If it already has changed once, just append to it ", Dates Added"
                                masterSheet.cell(row, 2).value = str(masterSheet.cell(row,2).value) + ", Dates Added"

                            # Paste the new dates in
                            masterSheet.cell(row, masterSheetCol).value = GTNSheet.cell(matchingRow, GTNCol).value
                            masterSheet.cell(row, masterSheetCol).fill = yellowFill

                        else:
                            # Else, if both are non-null, then paste it over
                            masterSheet.cell(row, masterSheetCol).fill = greenFill
                            masterSheet.cell(row, masterSheetCol).value = GTNSheet.cell(matchingRow, GTNCol).value

                    
                    # ELSE CASE: Anything that we don't need to compare, just paste and colour green
                    else:
                        masterSheet.cell(row, masterSheetCol).fill = greenFill
                        masterSheet.cell(row, masterSheetCol).value = GTNSheet.cell(matchingRow, GTNCol).value
                        


                # LOOP [for GTNCol in range(2, 52):] ENDS
                # GTN whole row has been pasted into the the mastersheet, proceed to delete it
                GTNSheet.delete_rows(matchingRow)

                # This breaks out of LOOP[for matchingRow in range(3, GTNSheet.max_row + 1):] since we can stop finding a match for this row in mastersheet 
                break

        # If DEBUG is ON, this will print out whether we found a match
        if DEBUG == 1:
            print("Match status: ", matchStatus)
            
        # If we can't find a match, matchStatus should remain as False, check just in case:
        if matchStatus == False:

            # If DEBUG is ON, this will print out which PO has been "Cancelled"
            if DEBUG == 1:
                print("PO Cancelled! PO: " , PO)


            # Update Exceptions Cell (masterSheetCol: 2) for this row to "Cancelled"
            masterSheet.cell(row, 2). value = "Cancelled"

            # Highlight the whole row as Red as per custom
            for fillCol in range(2, 52):
                masterSheet.cell(row, fillCol).fill = redFill

        # Formatting purpose for printed statements
        if DEBUG == 1:
            print("\n \n")

        # Newly added Aug-21-2020: If any of the columns from 48 to 51 contains the word "Milton", highlight the whole row
        miltonStatus = False
        for shipToCol in range(48, 52):
            if masterSheet.cell(row,shipToCol).value == None:
                continue
            elif "Milton" in masterSheet.cell(row, shipToCol).value or "milton" in masterSheet.cell(row, shipToCol).value:
                miltonStatus = True

        # If we find the word Milton in the cells, highlight the whole row yellow and add a decirption under "Exceptions" column
        if miltonStatus == True:
            for fillCol in range(2, 52):
                masterSheet.cell(row, fillCol).fill = yellowFill

            # First check if the "Exceptions" (MastersheetCol: 2) is empty
            if str(masterSheet.cell(row, 2).value).rstrip() == "None":

                # If it is, cell should be "Dates Added"
                masterSheet.cell(row, 2).value = "Milton Shipment"
            else:
                # If it already has changed once, just append to it ", Dates Added"
                masterSheet.cell(row, 2).value = str(masterSheet.cell(row,2).value) + ", Milton Shipment"
            
            
        

# Origin Lookup Fuction
# Purpose: Vlookup the Origin Offices for each factory
def lookUpOrigins(masterSheet, lookUpSheet):
    # Column is fixed to the factory column of 9
    originColumn = 4
    factoryColumn = 9

    # For each row, we have to check the "Factory Name" column
    for masterSheetRow in range(3, masterSheet.max_row + 1):
        origin = str(masterSheet.cell(masterSheetRow, originColumn).value).rstrip()
        factoryName = str(masterSheet.cell(masterSheetRow, factoryColumn).value).rstrip()

        if factoryName == "None":
            continue
        
        # Check the lookUpSheet if this factory exists
        for lookUpRow in range(2, lookUpSheet.max_row + 1):

            # If such factory exists
            if str(lookUpSheet.cell(lookUpRow, 1).value).rstrip() == factoryName:

                # Take the origin out, and loop is complete
                origin = str(lookUpSheet.cell(lookUpRow, 2).value).rstrip()
                
                break

        # So then the origin value in the masterSheet should then update the origin value
        masterSheet.cell(masterSheetRow, originColumn).value = origin            
                
        

                
# New Bookings Consolidation Main Function
# Purpose: Add the remaining unmatched GTNrows to the masterSheet, these are our new bookings.
def appendNewBookings(masterSheet, GTNSheet):

    # newBookingsRow starts from row = 2
    newBookingsRow = 2
    
    # Migrate the remaining rows in GTNSheet to mastersheet["New Bookings"]
    for row in range(3, GTNSheet.max_row + 1):
        
        # First, check if this current GTNrow is empty
        # To qualify as a null row, Origin and PO number and Destination fields are checked
        if str(GTNSheet.cell(row, 3).value).strip() == "None" or str(GTNSheet.cell(row, 6).value).strip() == "None" or str(GTNSheet.cell(row,36).value) == "None":
            # If this is a null row, skip this row.
            continue

        # For a non-empty row, first, update the "Exceptions" to denote as New Booking
        GTNSheet.cell(row, 1).value = "New Booking"
        
        # Paste each column of this current row to the last row of the masterSheet
        for col in range(1, 48):

            # If we're pasting origins
            if col == 3:
                # Take initial value first
                factory = str(GTNSheet.cell(row,8).value).rstrip()
                origin = str(GTNSheet.cell(row, col).value).rstrip()

                for i in range(2, lookUpSheet.max_row + 1):
                    if factory == str(lookUpSheet.cell(i,1).value).rstrip():
                        origin = str(lookUpSheet.cell(i,2).value).rstrip()
                        break              

                newBookingsSheet.cell(newBookingsRow, col).value = origin
            else:    
                newBookingsSheet.cell(newBookingsRow, col).value = GTNSheet.cell(row, col).value
                

            # Delete this cell value in GTNSheet
            GTNSheet.cell(row, col).value = ""
        
        # Go to next row in the new bookings sheet
        newBookingsRow += 1



def consolidateUpdates(masterSheet):
    lastEmptyUpdateRow = 2
    updateCount = 0
    for row in range(3, masterSheet.max_row + 1):
        # for each record in the mastersheet, find none-empty "exceptions"
        if str(masterSheet.cell(row, 2).value).strip() != "None" and str(masterSheet.cell(row,2).value).rstrip() != "YES" and str(masterSheet.cell(row, 2).value).strip() != "":
            updateCount += 1
            # Update HOD
            updateSheet.cell(lastEmptyUpdateRow, 1).value = masterSheet.cell(row, 14).value
            # Update Origin
            origin = str(masterSheet.cell(row, 4).value)
            updateSheet.cell(lastEmptyUpdateRow, 2).value = str(masterSheet.cell(row, 4).value)
            # Update Destination
            updateSheet.cell(lastEmptyUpdateRow, 3).value = str(masterSheet.cell(row, 37).value)
            # Update SO
            updateSheet.cell(lastEmptyUpdateRow, 4).value = str(masterSheet.cell(row, 7).value)
            # Update PO
            updateSheet.cell(lastEmptyUpdateRow, 5).value = str(masterSheet.cell(row, 15).value)
            # Update Line Item ID
            updateSheet.cell(lastEmptyUpdateRow, 6).value = str(masterSheet.cell(row, 18).value)
            # Update FLEX-ID
            updateSheet.cell(lastEmptyUpdateRow, 7).value = str(masterSheet.cell(row, 1).value)
            # Update Remarks
            updateSheet.cell(lastEmptyUpdateRow, 8).value = str(masterSheet.cell(row, 2).value)

            lastEmptyUpdateRow += 1

    return updateCount

def consolidateGTNReport(GTN, numberOfFiles):
    # clear the current GTN file
    for row in range(3, GTNSheet.max_row + 1):
        # place a cap on max rows to prevent overflow
        if row > 10000:
            break
        for col in range(1, 51):
            GTNSheet.cell(row,col).value = None

    # for empty sheets, last empty row must be 3rd row
    lastEmptyRow = 3
    
    # outside loop begins, for every single file
    for fileNumber in range(1, numberOfFiles + 1):
        # get the file name and add extension to it
        fileName = str(fileNumber) + ".xls"

        # open this file
        feederWb = xlrd.open_workbook(fileName)
        # get the first sheet in this workbook file
        feederSheet = feederWb.sheet_by_index(0)

        # assuming that we don't put an empty file, start pasting process
        for feederRow in range(1, feederSheet.nrows):
            if DEBUG == 1:
                print("Loop from 2 to ", feederSheet.nrows)
            for sharedCol in range (1, 50):
                if str(feederSheet.cell(feederRow, sharedCol).value).rstrip() == "None":
                    continue
                #Paste cell value
                if DEBUG == 1:
                    print("(row, col) : (", feederRow , "," , sharedCol, ")")
                    print("Cell value: " , feederSheet.cell(feederRow, sharedCol).value)
                GTNSheet.cell(lastEmptyRow, sharedCol + 1).value = feederSheet.cell(feederRow, sharedCol).value #sharedCol + 1 because xlrd starts from 0 and openpyxl starts from 1


            lastEmptyRow += 1

                
            

# Main
print("Please rename all feeder files according as [number].xlsx, starting from 1.xlsx.")
numberOfFiles = int(input("Please input number of feeder files: "))
print("Consolidating all feeder files into GTN report...")
consolidateGTNReport(GTNSheet, numberOfFiles)
print("Saving...")
GTNWb.save("GTN.xlsx")
print("Updating Mastersheet...")
updateExistingBookings(masterSheet, GTNSheet)
print("Mastersheet updated.")
print("Adding in new bookings...")
appendNewBookings(masterSheet, GTNSheet)
print("New bookings added.")
print("Looking up all origins...")
lookUpOrigins(masterSheet, lookUpSheet)
print("Origins looked up.")
print("Consolidating all updates...")
updateCount = consolidateUpdates(masterSheet)
print("Updates consolidated.")
masterSheetWb.save("Mastersheet_Output_" + datetime.datetime.now().strftime("%m%d%Y") + ".xlsx")
GTNWb.save("GTN_booking_Output_" + datetime.datetime.now().strftime("%m%d%Y") +".xlsx")

print("Job complete! Total number of updates: ", updateCount)
                    
                
                
                
                
        
    
    
    







































    
    
